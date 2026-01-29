"""
API views for multi-tenant: current user (me), CSRF token for session auth, and admin Excel upload.
"""
from datetime import date
from decimal import Decimal, InvalidOperation
import re

from django.middleware.csrf import get_token
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, IsAdminUser, AllowAny
from rest_framework.parsers import MultiPartParser, FormParser

from core.models import Organization, LeadChannel, UnitType, MilestonePhase, ComplaintCategory
from analytics.models import MarketingCampaign, LocationDemandMonthly, OrgKPI_Daily, ProjectKPI_Daily, OrgMonthlySnapshot
from projects.models import Project, Unit
from crm.models import Customer, Booking, CustomerSatisfactionSurvey, Complaint
from finance.models import Vendor, VendorBill, VendorPayment, CashFlowEntry
from construction.models import Contractor, Milestone, DailyProgress, DelayPenalty
from compliance.models import LegalCase, ComplianceItem, RERARegistration


class CsrfAPIView(APIView):
    """GET /api/csrf/ - Return CSRF token for session-authenticated POST (e.g. upload)."""
    permission_classes = [AllowAny]

    def get(self, request):
        return Response({"csrf": get_token(request)})


class MeAPIView(APIView):
    """
    GET /api/me/
    Returns current user and org for multi-tenant frontend.
    When authenticated with employee_profile, org_code is the user's org.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        user = request.user
        org_code = None
        is_staff = getattr(user, "is_staff", False)
        emp = getattr(user, "employee_profile", None)
        if emp and emp.organization:
            org_code = emp.organization.code
        return Response({
            "username": getattr(user, "username", None),
            "org_code": org_code,
            "is_staff": is_staff,
        })


def _parse_excel_upload(file, org):
    """
    Parse Excel file and create/update org-scoped data.
    Expected sheets:
    - Organization, Project, ProjectKPI_Daily, Unit, Customer, Booking
    - Vendor, VendorBill, VendorPayment, CashFlowEntry
    - Contractor, Milestone, DailyProgress, DelayPenalty
    - CustomerSatisfactionSurvey, Complaint
    - LegalCase, ComplianceItem, RERARegistration
    - OrgMonthlySnapshot, MarketingCampaign, LocationDemandMonthly, OrgKPI_Daily (optional)
    Returns (created_counts, errors).
    """
    try:
        import openpyxl
    except ImportError:
        return {}, ["openpyxl is required. Install: pip install openpyxl"]
    from openpyxl import load_workbook

    created = {
        "organizations": 0, "projects": 0, "project_kpi": 0, "units": 0, 
        "customers": 0, "bookings": 0, "vendors": 0, "vendor_bills": 0, 
        "vendor_payments": 0, "cashflow_entries": 0, "contractors": 0, 
        "milestones": 0, "daily_progress": 0, "delay_penalties": 0,
        "satisfaction_surveys": 0, "complaints": 0, "legal_cases": 0,
        "compliance_items": 0, "rera_registrations": 0,
        "marketing_campaigns": 0, "location_demand": 0, "org_kpi": 0, 
        "org_monthly_snapshot": 0
    }
    errors = []
    wb = load_workbook(file, read_only=True, data_only=True)

    def safe_decimal(value, field_name: str, row_preview) -> Decimal:
        """
        Convert Excel cell value to Decimal safely.
        - Strips currency symbols, commas and other non-numeric characters.
        - Returns Decimal(0) on failure and records a soft warning.
        """
        if value is None or value == "":
            return Decimal(0)
        s = str(value)
        # keep digits, dot and minus
        s = re.sub(r"[^\d\.\-]", "", s)
        if not s:
            return Decimal(0)
        try:
            return Decimal(s)
        except InvalidOperation:
            errors.append(f"Invalid decimal for {field_name} in row {row_preview}: {value!r}")
            return Decimal(0)

    def safe_date(value, field_name: str, row_preview):
        """Convert Excel cell value to date safely. Returns None on failure."""
        if value is None or value == "":
            return None
        if hasattr(value, "date"):
            return value.date()
        if isinstance(value, str):
            try:
                return date.fromisoformat(value)
            except:
                pass
        return None

    # Sheet "Organization": optional; if present, ensure org exists (code must match request org_code if given)
    if "Organization" in wb.sheetnames:
        ws = wb["Organization"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or row[0] is None:
                continue
            code = str(row[0]).strip() if row[0] else None
            name = str(row[1]).strip() if len(row) > 1 and row[1] else (code or "")
            if not code:
                continue
            if org and code != org.code:
                errors.append(f"Organization sheet code '{code}' does not match upload org_code '{org.code}'.")
                break
            if not org:
                org, _ = Organization.objects.get_or_create(code=code, defaults={"name": name or code})
                created["organizations"] += 1
            break
    if not org:
        errors.append("Provide org_code in request or Organization sheet with code, name.")
        return created, errors

    # Sheet "Project"
    if "Project" in wb.sheetnames:
        ws = wb["Project"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or row[0] is None:
                continue
            try:
                # Expected columns:
                # 0: project_code (required)
                # 1: name (required)
                # 2: location (required)
                # 3: city (optional)
                # 4: status (optional: on_track, at_risk, delayed, stalled, completed)
                # 5: planned_start_date (optional)
                # 6: actual_start_date (optional)
                # 7: planned_completion_date (optional)
                # 8: expected_completion_date (optional)
                # 9: actual_completion_date (optional)
                # 10: budget (optional)
                # 11: rera_registration_number (optional)
                # 12: rera_valid_until (optional)
                project_code = str(row[0]).strip() if row[0] else None
                if not project_code:
                    continue
                
                name = str(row[1]).strip() if len(row) > 1 and row[1] else project_code
                location = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                city = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                
                # Status
                status = "on_track"
                if len(row) > 4 and row[4]:
                    s = str(row[4]).strip().lower()
                    if s in ("on_track", "at_risk", "delayed", "stalled", "completed"):
                        status = s
                
                # Dates
                planned_start_date = safe_date(row[5] if len(row) > 5 else None, "planned_start_date", row[:2])
                actual_start_date = safe_date(row[6] if len(row) > 6 else None, "actual_start_date", row[:2])
                planned_completion_date = safe_date(row[7] if len(row) > 7 else None, "planned_completion_date", row[:2])
                expected_completion_date = safe_date(row[8] if len(row) > 8 else None, "expected_completion_date", row[:2])
                actual_completion_date = safe_date(row[9] if len(row) > 9 else None, "actual_completion_date", row[:2])
                
                # Budget
                budget = safe_decimal(row[10] if len(row) > 10 else None, "budget", row[:2])
                
                # RERA
                rera_registration_number = str(row[11]).strip() if len(row) > 11 and row[11] else ""
                rera_valid_until = safe_date(row[12] if len(row) > 12 else None, "rera_valid_until", row[:2])
                
                Project.objects.update_or_create(
                    organization=org, project_code=project_code,
                    defaults={
                        "name": name,
                        "location": location,
                        "city": city,
                        "status": status,
                        "planned_start_date": planned_start_date,
                        "actual_start_date": actual_start_date,
                        "planned_completion_date": planned_completion_date,
                        "expected_completion_date": expected_completion_date,
                        "actual_completion_date": actual_completion_date,
                        "budget": budget,
                        "rera_registration_number": rera_registration_number,
                        "rera_valid_until": rera_valid_until,
                    }
                )
                created["projects"] += 1
            except Exception as e:
                errors.append(f"Project row {row[:2]}: {e}")

    # Sheet "ProjectKPI_Daily"
    if "ProjectKPI_Daily" in wb.sheetnames:
        ws = wb["ProjectKPI_Daily"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or row[0] is None:
                continue
            try:
                # Expected columns:
                # 0: project_code (required) - matches Project.project_code
                # 1: date (required)
                # 2: total_units (optional, default 0)
                # 3: sold_units (optional, default 0)
                # 4: booked_units (optional, default 0)
                # 5: blocked_units (optional, default 0)
                # 6: unsold_units (optional, default 0)
                # 7: revenue_booked (optional, default 0)
                # 8: revenue_collected (optional, default 0)
                # 9: outstanding (optional, default 0)
                # 10: construction_percent (optional, default 0)
                # 11: satisfaction_avg (optional, default 0)
                # 12: budget (optional, default 0)
                # 13: cost_incurred (optional, default 0)
                # 14: margin_percent (optional, default 0)
                project_code = str(row[0]).strip() if row[0] else None
                if not project_code:
                    continue
                
                project = Project.objects.filter(organization=org, project_code=project_code).first()
                if not project:
                    errors.append(f"ProjectKPI_Daily row {row[:2]}: Project '{project_code}' not found")
                    continue
                
                kpi_date = safe_date(row[1] if len(row) > 1 else None, "date", row[:2])
                if not kpi_date:
                    errors.append(f"ProjectKPI_Daily row {row[:2]}: Invalid or missing date")
                    continue
                
                total_units = int(row[2]) if len(row) > 2 and row[2] is not None else 0
                sold_units = int(row[3]) if len(row) > 3 and row[3] is not None else 0
                booked_units = int(row[4]) if len(row) > 4 and row[4] is not None else 0
                blocked_units = int(row[5]) if len(row) > 5 and row[5] is not None else 0
                unsold_units = int(row[6]) if len(row) > 6 and row[6] is not None else 0
                
                revenue_booked = safe_decimal(row[7] if len(row) > 7 else None, "revenue_booked", row[:2])
                revenue_collected = safe_decimal(row[8] if len(row) > 8 else None, "revenue_collected", row[:2])
                outstanding = safe_decimal(row[9] if len(row) > 9 else None, "outstanding", row[:2])
                
                construction_percent = safe_decimal(row[10] if len(row) > 10 else None, "construction_percent", row[:2])
                satisfaction_avg = safe_decimal(row[11] if len(row) > 11 else None, "satisfaction_avg", row[:2])
                
                budget = safe_decimal(row[12] if len(row) > 12 else None, "budget", row[:2])
                cost_incurred = safe_decimal(row[13] if len(row) > 13 else None, "cost_incurred", row[:2])
                margin_percent = safe_decimal(row[14] if len(row) > 14 else None, "margin_percent", row[:2])
                
                ProjectKPI_Daily.objects.update_or_create(
                    project=project, date=kpi_date,
                    defaults={
                        "total_units": total_units,
                        "sold_units": sold_units,
                        "booked_units": booked_units,
                        "blocked_units": blocked_units,
                        "unsold_units": unsold_units,
                        "revenue_booked": revenue_booked,
                        "revenue_collected": revenue_collected,
                        "outstanding": outstanding,
                        "construction_percent": construction_percent,
                        "satisfaction_avg": satisfaction_avg,
                        "budget": budget,
                        "cost_incurred": cost_incurred,
                        "margin_percent": margin_percent,
                    }
                )
                created["project_kpi"] += 1
            except Exception as e:
                errors.append(f"ProjectKPI_Daily row {row[:2]}: {e}")

    # Sheet "Unit"
    if "Unit" in wb.sheetnames:
        ws = wb["Unit"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or row[0] is None:
                continue
            try:
                project_code = str(row[0]).strip() if row[0] else None
                if not project_code:
                    continue
                project = Project.objects.filter(organization=org, project_code=project_code).first()
                if not project:
                    errors.append(f"Unit row {row[:2]}: Project '{project_code}' not found")
                    continue
                unit_number = str(row[1]).strip() if len(row) > 1 and row[1] else None
                if not unit_number:
                    continue
                unit_type = None
                if len(row) > 2 and row[2]:
                    unit_type_code = str(row[2]).strip()
                    unit_type = UnitType.objects.filter(organization=org, code=unit_type_code).first()
                    if not unit_type:
                        unit_type, _ = UnitType.objects.get_or_create(
                            organization=org, code=unit_type_code,
                            defaults={"label": unit_type_code.upper(), "is_active": True}
                        )
                floor = int(row[3]) if len(row) > 3 and row[3] is not None else 0
                tower = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                carpet_area = safe_decimal(row[5] if len(row) > 5 else None, "carpet_area", row[:2])
                built_up_area = safe_decimal(row[6] if len(row) > 6 else None, "built_up_area", row[:2])
                base_price = safe_decimal(row[7] if len(row) > 7 else None, "base_price", row[:2])
                final_price = safe_decimal(row[8] if len(row) > 8 else None, "final_price", row[:2])
                status = "available"
                if len(row) > 9 and row[9]:
                    s = str(row[9]).strip().lower()
                    if s in ("available", "blocked", "booked", "sold"):
                        status = s
                listed_date = safe_date(row[10] if len(row) > 10 else None, "listed_date", row[:2])
                Unit.objects.update_or_create(
                    project=project, unit_number=unit_number,
                    defaults={
                        "unit_type": unit_type, "floor": floor, "tower": tower,
                        "carpet_area": carpet_area, "built_up_area": built_up_area,
                        "base_price": base_price, "final_price": final_price,
                        "status": status, "listed_date": listed_date,
                    }
                )
                created["units"] += 1
            except Exception as e:
                errors.append(f"Unit row {row[:2]}: {e}")

    # Sheet "Customer"
    if "Customer" in wb.sheetnames:
        ws = wb["Customer"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or row[0] is None:
                continue
            try:
                customer_code = str(row[0]).strip() if row[0] else None
                if not customer_code:
                    continue
                name = str(row[1]).strip() if len(row) > 1 and row[1] else customer_code
                email = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                phone = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                if not phone:
                    errors.append(f"Customer row {row[:2]}: Missing phone")
                    continue
                project = None
                if len(row) > 4 and row[4]:
                    project_code = str(row[4]).strip()
                    project = Project.objects.filter(organization=org, project_code=project_code).first()
                unit = None
                if project and len(row) > 5 and row[5]:
                    unit_number = str(row[5]).strip()
                    unit = Unit.objects.filter(project=project, unit_number=unit_number).first()
                customer_channel = channel
                if len(row) > 6 and row[6]:
                    channel_code = str(row[6]).strip()
                    customer_channel = LeadChannel.objects.filter(organization=org, code=channel_code).first() or channel
                status = "walk_in"
                if len(row) > 7 and row[7]:
                    s = str(row[7]).strip().lower()
                    if s in ("walk_in", "applied", "booked", "possession", "cancelled"):
                        status = s
                walk_in_date = safe_date(row[8] if len(row) > 8 else None, "walk_in_date", row[:2])
                application_date = safe_date(row[9] if len(row) > 9 else None, "application_date", row[:2])
                booking_date = safe_date(row[10] if len(row) > 10 else None, "booking_date", row[:2])
                possession_date = safe_date(row[11] if len(row) > 11 else None, "possession_date", row[:2])
                cancellation_date = safe_date(row[12] if len(row) > 12 else None, "cancellation_date", row[:2])
                satisfaction_score_cached = safe_decimal(row[13] if len(row) > 13 else None, "satisfaction_score_cached", row[:2])
                if satisfaction_score_cached > Decimal(5):
                    satisfaction_score_cached = Decimal(5)
                if satisfaction_score_cached < Decimal(0):
                    satisfaction_score_cached = Decimal(0)
                Customer.objects.update_or_create(
                    organization=org, customer_code=customer_code,
                    defaults={
                        "name": name, "email": email, "phone": phone,
                        "project": project, "unit": unit, "channel": customer_channel,
                        "status": status, "walk_in_date": walk_in_date,
                        "application_date": application_date, "booking_date": booking_date,
                        "possession_date": possession_date, "cancellation_date": cancellation_date,
                        "satisfaction_score_cached": satisfaction_score_cached,
                    }
                )
                created["customers"] += 1
            except Exception as e:
                errors.append(f"Customer row {row[:2]}: {e}")

    # Sheet "Booking"
    if "Booking" in wb.sheetnames:
        ws = wb["Booking"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or row[0] is None:
                continue
            try:
                customer_code = str(row[0]).strip() if row[0] else None
                if not customer_code:
                    continue
                customer = Customer.objects.filter(organization=org, customer_code=customer_code).first()
                if not customer:
                    errors.append(f"Booking row {row[:2]}: Customer '{customer_code}' not found")
                    continue
                project_code = str(row[1]).strip() if len(row) > 1 and row[1] else None
                if not project_code:
                    errors.append(f"Booking row {row[:2]}: Missing project_code")
                    continue
                project = Project.objects.filter(organization=org, project_code=project_code).first()
                if not project:
                    errors.append(f"Booking row {row[:2]}: Project '{project_code}' not found")
                    continue
                unit_number = str(row[2]).strip() if len(row) > 2 and row[2] else None
                if not unit_number:
                    errors.append(f"Booking row {row[:2]}: Missing unit_number")
                    continue
                unit = Unit.objects.filter(project=project, unit_number=unit_number).first()
                if not unit:
                    errors.append(f"Booking row {row[:2]}: Unit '{unit_number}' not found")
                    continue
                booking_value = safe_decimal(row[3] if len(row) > 3 else None, "booking_value", row[:2])
                if booking_value == Decimal(0):
                    errors.append(f"Booking row {row[:2]}: booking_value is required")
                    continue
                booking_date = safe_date(row[4] if len(row) > 4 else None, "booking_date", row[:2])
                if not booking_date:
                    errors.append(f"Booking row {row[:2]}: Invalid booking_date")
                    continue
                status = "active"
                if len(row) > 5 and row[5]:
                    s = str(row[5]).strip().lower()
                    if s in ("active", "cancelled"):
                        status = s
                Booking.objects.update_or_create(
                    customer=customer, project=project, unit=unit, booking_date=booking_date,
                    defaults={"booking_value": booking_value, "status": status}
                )
                created["bookings"] += 1
            except Exception as e:
                errors.append(f"Booking row {row[:2]}: {e}")

    # Sheet "Vendor"
    if "Vendor" in wb.sheetnames:
        ws = wb["Vendor"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or row[0] is None:
                continue
            try:
                name = str(row[0]).strip()
                if not name:
                    continue
                Vendor.objects.get_or_create(organization=org, name=name)
                created["vendors"] += 1
            except Exception as e:
                errors.append(f"Vendor row {row[:2]}: {e}")

    # Sheet "VendorBill"
    if "VendorBill" in wb.sheetnames:
        ws = wb["VendorBill"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or row[0] is None:
                continue
            try:
                vendor_name = str(row[0]).strip() if row[0] else None
                if not vendor_name:
                    continue
                vendor = Vendor.objects.filter(organization=org, name=vendor_name).first()
                if not vendor:
                    errors.append(f"VendorBill row {row[:2]}: Vendor '{vendor_name}' not found")
                    continue
                project_code = str(row[1]).strip() if len(row) > 1 and row[1] else None
                if not project_code:
                    errors.append(f"VendorBill row {row[:2]}: Missing project_code")
                    continue
                project = Project.objects.filter(organization=org, project_code=project_code).first()
                if not project:
                    errors.append(f"VendorBill row {row[:2]}: Project '{project_code}' not found")
                    continue
                bill_no = str(row[2]).strip() if len(row) > 2 and row[2] else None
                if not bill_no:
                    errors.append(f"VendorBill row {row[:2]}: Missing bill_no")
                    continue
                bill_date = safe_date(row[3] if len(row) > 3 else None, "bill_date", row[:2])
                if not bill_date:
                    errors.append(f"VendorBill row {row[:2]}: Invalid bill_date")
                    continue
                due_date = safe_date(row[4] if len(row) > 4 else None, "due_date", row[:2])
                amount = safe_decimal(row[5] if len(row) > 5 else None, "amount", row[:2])
                if amount == Decimal(0):
                    errors.append(f"VendorBill row {row[:2]}: amount is required")
                    continue
                status = "unpaid"
                if len(row) > 6 and row[6]:
                    s = str(row[6]).strip().lower()
                    if s in ("unpaid", "partial", "paid"):
                        status = s
                VendorBill.objects.update_or_create(
                    vendor=vendor, bill_no=bill_no,
                    defaults={
                        "project": project, "bill_date": bill_date,
                        "due_date": due_date, "amount": amount, "status": status
                    }
                )
                created["vendor_bills"] += 1
            except Exception as e:
                errors.append(f"VendorBill row {row[:2]}: {e}")

    # Sheet "VendorPayment"
    if "VendorPayment" in wb.sheetnames:
        ws = wb["VendorPayment"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or row[0] is None:
                continue
            try:
                vendor_name = str(row[0]).strip() if row[0] else None
                bill_no = str(row[1]).strip() if len(row) > 1 and row[1] else None
                if not vendor_name or not bill_no:
                    continue
                vendor = Vendor.objects.filter(organization=org, name=vendor_name).first()
                if not vendor:
                    errors.append(f"VendorPayment row {row[:2]}: Vendor '{vendor_name}' not found")
                    continue
                bill = VendorBill.objects.filter(vendor=vendor, bill_no=bill_no).first()
                if not bill:
                    errors.append(f"VendorPayment row {row[:2]}: Bill '{bill_no}' not found")
                    continue
                amount = safe_decimal(row[2] if len(row) > 2 else None, "amount", row[:2])
                if amount == Decimal(0):
                    errors.append(f"VendorPayment row {row[:2]}: amount is required")
                    continue
                paid_on = safe_date(row[3] if len(row) > 3 else None, "paid_on", row[:2])
                if not paid_on:
                    errors.append(f"VendorPayment row {row[:2]}: Invalid paid_on date")
                    continue
                reference = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                VendorPayment.objects.create(
                    bill=bill, amount=amount, paid_on=paid_on, reference=reference
                )
                created["vendor_payments"] += 1
            except Exception as e:
                errors.append(f"VendorPayment row {row[:2]}: {e}")

    # Sheet "CashFlowEntry"
    if "CashFlowEntry" in wb.sheetnames:
        ws = wb["CashFlowEntry"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or row[0] is None:
                continue
            try:
                flow_type_str = str(row[0]).strip().lower() if row[0] else None
                if not flow_type_str or flow_type_str not in ("inflow", "outflow"):
                    errors.append(f"CashFlowEntry row {row[:2]}: flow_type must be 'inflow' or 'outflow'")
                    continue
                flow_type = "inflow" if flow_type_str == "inflow" else "outflow"
                amount = safe_decimal(row[1] if len(row) > 1 else None, "amount", row[:2])
                if amount == Decimal(0):
                    errors.append(f"CashFlowEntry row {row[:2]}: amount is required")
                    continue
                entry_date = safe_date(row[2] if len(row) > 2 else None, "date", row[:2])
                if not entry_date:
                    errors.append(f"CashFlowEntry row {row[:2]}: Invalid date")
                    continue
                project_code = str(row[3]).strip() if len(row) > 3 and row[3] else None
                project = None
                if project_code:
                    project = Project.objects.filter(organization=org, project_code=project_code).first()
                category = str(row[4]).strip() if len(row) > 4 and row[4] else ""
                description = str(row[5]).strip() if len(row) > 5 and row[5] else ""
                CashFlowEntry.objects.create(
                    organization=org, project=project, flow_type=flow_type,
                    amount=amount, date=entry_date, category=category, description=description
                )
                created["cashflow_entries"] += 1
            except Exception as e:
                errors.append(f"CashFlowEntry row {row[:2]}: {e}")

    # Sheet "Contractor"
    if "Contractor" in wb.sheetnames:
        ws = wb["Contractor"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or row[0] is None:
                continue
            try:
                name = str(row[0]).strip()
                if not name:
                    continue
                specialization = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                performance_score = safe_decimal(row[2] if len(row) > 2 else None, "performance_score", row[:2])
                if performance_score > Decimal(100):
                    performance_score = Decimal(100)
                if performance_score < Decimal(0):
                    performance_score = Decimal(0)
                is_active = True
                if len(row) > 3 and row[3] is not None:
                    is_active = bool(row[3])
                Contractor.objects.update_or_create(
                    organization=org, name=name,
                    defaults={
                        "specialization": specialization,
                        "performance_score": performance_score,
                        "is_active": is_active
                    }
                )
                created["contractors"] += 1
            except Exception as e:
                errors.append(f"Contractor row {row[:2]}: {e}")

    # Sheet "Milestone"
    if "Milestone" in wb.sheetnames:
        ws = wb["Milestone"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or row[0] is None:
                continue
            try:
                project_code = str(row[0]).strip() if row[0] else None
                if not project_code:
                    continue
                project = Project.objects.filter(organization=org, project_code=project_code).first()
                if not project:
                    errors.append(f"Milestone row {row[:2]}: Project '{project_code}' not found")
                    continue
                name = str(row[1]).strip() if len(row) > 1 and row[1] else None
                if not name:
                    errors.append(f"Milestone row {row[:2]}: Missing name")
                    continue
                planned_start = safe_date(row[2] if len(row) > 2 else None, "planned_start", row[:2])
                if not planned_start:
                    errors.append(f"Milestone row {row[:2]}: Invalid planned_start")
                    continue
                planned_end = safe_date(row[3] if len(row) > 3 else None, "planned_end", row[:2])
                if not planned_end:
                    errors.append(f"Milestone row {row[:2]}: Invalid planned_end")
                    continue
                actual_start = safe_date(row[4] if len(row) > 4 else None, "actual_start", row[:2])
                actual_end = safe_date(row[5] if len(row) > 5 else None, "actual_end", row[:2])
                status = "not_started"
                if len(row) > 6 and row[6]:
                    s = str(row[6]).strip().lower()
                    if s in ("not_started", "in_progress", "completed", "delayed"):
                        status = s
                completion_percent = safe_decimal(row[7] if len(row) > 7 else None, "completion_percent", row[:2])
                if completion_percent > Decimal(100):
                    completion_percent = Decimal(100)
                if completion_percent < Decimal(0):
                    completion_percent = Decimal(0)
                contractor_name = str(row[8]).strip() if len(row) > 8 and row[8] else None
                contractor = None
                if contractor_name:
                    contractor = Contractor.objects.filter(organization=org, name=contractor_name).first()
                contractor_score = None
                if len(row) > 9 and row[9] is not None:
                    contractor_score = safe_decimal(row[9], "contractor_score", row[:2])
                    if contractor_score > Decimal(10):
                        contractor_score = Decimal(10)
                    if contractor_score < Decimal(0):
                        contractor_score = Decimal(0)
                order = int(row[10]) if len(row) > 10 and row[10] is not None else 0
                phase_code = str(row[11]).strip() if len(row) > 11 and row[11] else None
                phase = None
                if phase_code:
                    phase = MilestonePhase.objects.filter(organization=org, code=phase_code).first()
                Milestone.objects.update_or_create(
                    project=project, name=name,
                    defaults={
                        "phase": phase, "planned_start": planned_start, "planned_end": planned_end,
                        "actual_start": actual_start, "actual_end": actual_end, "status": status,
                        "completion_percent": completion_percent, "contractor": contractor,
                        "contractor_score": contractor_score, "order": order
                    }
                )
                created["milestones"] += 1
            except Exception as e:
                errors.append(f"Milestone row {row[:2]}: {e}")

    # Sheet "DailyProgress"
    if "DailyProgress" in wb.sheetnames:
        ws = wb["DailyProgress"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or row[0] is None:
                continue
            try:
                project_code = str(row[0]).strip() if row[0] else None
                if not project_code:
                    continue
                project = Project.objects.filter(organization=org, project_code=project_code).first()
                if not project:
                    errors.append(f"DailyProgress row {row[:2]}: Project '{project_code}' not found")
                    continue
                progress_date = safe_date(row[1] if len(row) > 1 else None, "date", row[:2])
                if not progress_date:
                    errors.append(f"DailyProgress row {row[:2]}: Invalid date")
                    continue
                planned_percent = safe_decimal(row[2] if len(row) > 2 else None, "planned_percent", row[:2])
                actual_percent = safe_decimal(row[3] if len(row) > 3 else None, "actual_percent", row[:2])
                workers_present = int(row[4]) if len(row) > 4 and row[4] is not None else 0
                equipment_deployed = int(row[5]) if len(row) > 5 and row[5] is not None else 0
                notes = str(row[6]).strip() if len(row) > 6 and row[6] else ""
                DailyProgress.objects.update_or_create(
                    project=project, date=progress_date,
                    defaults={
                        "planned_percent": planned_percent, "actual_percent": actual_percent,
                        "workers_present": workers_present, "equipment_deployed": equipment_deployed,
                        "notes": notes
                    }
                )
                created["daily_progress"] += 1
            except Exception as e:
                errors.append(f"DailyProgress row {row[:2]}: {e}")

    # Sheet "DelayPenalty"
    if "DelayPenalty" in wb.sheetnames:
        ws = wb["DelayPenalty"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or row[0] is None:
                continue
            try:
                project_code = str(row[0]).strip() if row[0] else None
                if not project_code:
                    continue
                project = Project.objects.filter(organization=org, project_code=project_code).first()
                if not project:
                    errors.append(f"DelayPenalty row {row[:2]}: Project '{project_code}' not found")
                    continue
                milestone_name = str(row[1]).strip() if len(row) > 1 and row[1] else None
                milestone = None
                if milestone_name:
                    milestone = Milestone.objects.filter(project=project, name=milestone_name).first()
                contractor_name = str(row[2]).strip() if len(row) > 2 and row[2] else None
                contractor = None
                if contractor_name:
                    contractor = Contractor.objects.filter(organization=org, name=contractor_name).first()
                delay_days = int(row[3]) if len(row) > 3 and row[3] is not None else 0
                penalty_per_day = safe_decimal(row[4] if len(row) > 4 else None, "penalty_per_day", row[:2])
                penalty_amount = safe_decimal(row[5] if len(row) > 5 else None, "penalty_amount", row[:2])
                pending_recovery = safe_decimal(row[6] if len(row) > 6 else None, "pending_recovery", row[:2])
                critical_escalations = int(row[7]) if len(row) > 7 and row[7] is not None else 0
                escalation_level = "low"
                if len(row) > 8 and row[8]:
                    s = str(row[8]).strip().lower()
                    if s in ("low", "medium", "high", "critical"):
                        escalation_level = s
                recorded_on = safe_date(row[9] if len(row) > 9 else None, "recorded_on", row[:2])
                if not recorded_on:
                    recorded_on = date.today()
                DelayPenalty.objects.create(
                    project=project, milestone=milestone, contractor=contractor,
                    delay_days=delay_days, penalty_per_day=penalty_per_day,
                    penalty_amount=penalty_amount, pending_recovery=pending_recovery,
                    critical_escalations=critical_escalations, escalation_level=escalation_level,
                    recorded_on=recorded_on
                )
                created["delay_penalties"] += 1
            except Exception as e:
                errors.append(f"DelayPenalty row {row[:2]}: {e}")

    # Sheet "CustomerSatisfactionSurvey"
    if "CustomerSatisfactionSurvey" in wb.sheetnames:
        ws = wb["CustomerSatisfactionSurvey"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or row[0] is None:
                continue
            try:
                customer_code = str(row[0]).strip() if row[0] else None
                if not customer_code:
                    continue
                customer = Customer.objects.filter(organization=org, customer_code=customer_code).first()
                if not customer:
                    errors.append(f"CustomerSatisfactionSurvey row {row[:2]}: Customer '{customer_code}' not found")
                    continue
                project_code = str(row[1]).strip() if len(row) > 1 and row[1] else None
                if not project_code:
                    errors.append(f"CustomerSatisfactionSurvey row {row[:2]}: Missing project_code")
                    continue
                project = Project.objects.filter(organization=org, project_code=project_code).first()
                if not project:
                    errors.append(f"CustomerSatisfactionSurvey row {row[:2]}: Project '{project_code}' not found")
                    continue
                score = safe_decimal(row[2] if len(row) > 2 else None, "score", row[:2])
                if score > Decimal(5):
                    score = Decimal(5)
                if score < Decimal(0):
                    score = Decimal(0)
                feedback = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                CustomerSatisfactionSurvey.objects.create(
                    organization=org, customer=customer, project=project,
                    score=score, feedback=feedback
                )
                created["satisfaction_surveys"] += 1
            except Exception as e:
                errors.append(f"CustomerSatisfactionSurvey row {row[:2]}: {e}")

    # Sheet "Complaint"
    if "Complaint" in wb.sheetnames:
        ws = wb["Complaint"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or row[0] is None:
                continue
            try:
                customer_code = str(row[0]).strip() if row[0] else None
                if not customer_code:
                    continue
                customer = Customer.objects.filter(organization=org, customer_code=customer_code).first()
                if not customer:
                    errors.append(f"Complaint row {row[:2]}: Customer '{customer_code}' not found")
                    continue
                project_code = str(row[1]).strip() if len(row) > 1 and row[1] else None
                if not project_code:
                    errors.append(f"Complaint row {row[:2]}: Missing project_code")
                    continue
                project = Project.objects.filter(organization=org, project_code=project_code).first()
                if not project:
                    errors.append(f"Complaint row {row[:2]}: Project '{project_code}' not found")
                    continue
                description = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                if not description:
                    errors.append(f"Complaint row {row[:2]}: Missing description")
                    continue
                category_code = str(row[3]).strip() if len(row) > 3 and row[3] else None
                category = None
                if category_code:
                    category = ComplaintCategory.objects.filter(organization=org, code=category_code).first()
                status = "open"
                if len(row) > 4 and row[4]:
                    s = str(row[4]).strip().lower()
                    if s in ("open", "in_progress", "resolved", "escalated"):
                        status = s
                risk_score = safe_decimal(row[5] if len(row) > 5 else None, "risk_score", row[:2])
                if risk_score > Decimal(100):
                    risk_score = Decimal(100)
                if risk_score < Decimal(0):
                    risk_score = Decimal(0)
                Complaint.objects.create(
                    customer=customer, project=project, category=category,
                    description=description, status=status, risk_score=risk_score
                )
                created["complaints"] += 1
            except Exception as e:
                errors.append(f"Complaint row {row[:2]}: {e}")

    # Sheet "LegalCase"
    if "LegalCase" in wb.sheetnames:
        ws = wb["LegalCase"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or row[0] is None:
                continue
            try:
                project_code = str(row[0]).strip() if row[0] else None
                if not project_code:
                    continue
                project = Project.objects.filter(organization=org, project_code=project_code).first()
                if not project:
                    errors.append(f"LegalCase row {row[:2]}: Project '{project_code}' not found")
                    continue
                case_id = str(row[1]).strip() if len(row) > 1 and row[1] else None
                if not case_id:
                    errors.append(f"LegalCase row {row[:2]}: Missing case_id")
                    continue
                case_type = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                description = str(row[3]).strip() if len(row) > 3 and row[3] else ""
                severity = str(row[4]).strip().lower() if len(row) > 4 and row[4] else "medium"
                status_str = str(row[5]).strip().lower() if len(row) > 5 and row[5] else "pending"
                filing_date = safe_date(row[6] if len(row) > 6 else None, "filing_date", row[:2])
                if not filing_date:
                    filing_date = date.today()
                LegalCase.objects.update_or_create(
                    project=project, case_id=case_id,
                    defaults={
                        "case_type": case_type, "description": description,
                        "severity": severity, "status": status_str, "filing_date": filing_date
                    }
                )
                created["legal_cases"] += 1
            except Exception as e:
                errors.append(f"LegalCase row {row[:2]}: {e}")

    # Sheet "ComplianceItem"
    if "ComplianceItem" in wb.sheetnames:
        ws = wb["ComplianceItem"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or row[0] is None:
                continue
            try:
                project_code = str(row[0]).strip() if row[0] else None
                if not project_code:
                    continue
                project = Project.objects.filter(organization=org, project_code=project_code).first()
                if not project:
                    errors.append(f"ComplianceItem row {row[:2]}: Project '{project_code}' not found")
                    continue
                item_name = str(row[1]).strip() if len(row) > 1 and row[1] else None
                if not item_name:
                    errors.append(f"ComplianceItem row {row[:2]}: Missing item_name")
                    continue
                description = str(row[2]).strip() if len(row) > 2 and row[2] else ""
                status = "pending"
                if len(row) > 3 and row[3]:
                    s = str(row[3]).strip().lower()
                    if s in ("compliant", "pending", "non_compliant"):
                        status = s
                due_date = safe_date(row[4] if len(row) > 4 else None, "due_date", row[:2])
                completed_date = safe_date(row[5] if len(row) > 5 else None, "completed_date", row[:2])
                ComplianceItem.objects.update_or_create(
                    project=project, item_name=item_name,
                    defaults={
                        "description": description, "status": status,
                        "due_date": due_date, "completed_date": completed_date
                    }
                )
                created["compliance_items"] += 1
            except Exception as e:
                errors.append(f"ComplianceItem row {row[:2]}: {e}")

    # Sheet "RERARegistration"
    if "RERARegistration" in wb.sheetnames:
        ws = wb["RERARegistration"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or row[0] is None:
                continue
            try:
                project_code = str(row[0]).strip() if row[0] else None
                if not project_code:
                    continue
                project = Project.objects.filter(organization=org, project_code=project_code).first()
                if not project:
                    errors.append(f"RERARegistration row {row[:2]}: Project '{project_code}' not found")
                    continue
                registration_number = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                if not registration_number:
                    errors.append(f"RERARegistration row {row[:2]}: Missing registration_number")
                    continue
                valid_until = safe_date(row[2] if len(row) > 2 else None, "valid_until", row[:2])
                if not valid_until:
                    errors.append(f"RERARegistration row {row[:2]}: Invalid valid_until")
                    continue
                status = "pending"
                if len(row) > 3 and row[3]:
                    s = str(row[3]).strip().lower()
                    if s in ("compliant", "pending", "expired"):
                        status = s
                RERARegistration.objects.update_or_create(
                    project=project,
                    defaults={
                        "registration_number": registration_number,
                        "valid_until": valid_until,
                        "status": status
                    }
                )
                created["rera_registrations"] += 1
            except Exception as e:
                errors.append(f"RERARegistration row {row[:2]}: {e}")

    # Ensure at least one LeadChannel for marketing
    channel = LeadChannel.objects.filter(organization=org).first()
    if not channel:
        channel, _ = LeadChannel.objects.get_or_create(
            organization=org, code="DIGITAL", defaults={"label": "Digital", "is_active": True}
        )

    # Sheet "MarketingCampaign"
    if "MarketingCampaign" in wb.sheetnames:
        ws = wb["MarketingCampaign"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or row[0] is None:
                continue
            try:
                name = str(row[0]).strip()
                campaign_code = str(row[1]).strip() if len(row) > 1 and row[1] else f"{org.code}-M-{row[0]}"
                # Expected columns:
                # 0: name
                # 1: campaign_code
                # 2: channel_code (optional, currently ignored / mapped to default channel)
                # 3: start_date
                # 4: end_date
                # 5: spend
                # 6: leads
                # 7: bookings
                # 8: cost_per_lead
                # 9: cost_per_booking
                # 10: roi
                # 11: status
                start_date = row[3] if len(row) > 3 else date.today()
                end_date = row[4] if len(row) > 4 else date.today()
                if hasattr(start_date, "date"):
                    start_date = start_date.date()
                if hasattr(end_date, "date"):
                    end_date = end_date.date()
                spend = safe_decimal(row[5] if len(row) > 5 else None, "spend", row[:2])
                leads = int(row[6]) if len(row) > 6 and row[6] is not None else 0
                bookings = int(row[7]) if len(row) > 7 and row[7] is not None else 0
                cost_per_lead = safe_decimal(row[8] if len(row) > 8 else None, "cost_per_lead", row[:2])
                cost_per_booking = safe_decimal(row[9] if len(row) > 9 else None, "cost_per_booking", row[:2])
                roi = safe_decimal(row[10] if len(row) > 10 else None, "roi", row[:2])
                status = "on_track"
                if len(row) > 11 and row[11]:
                    s = str(row[11]).strip().lower()
                    if s in ("on_track", "at_risk", "paused", "completed"):
                        status = s
                MarketingCampaign.objects.update_or_create(
                    organization=org, campaign_code=campaign_code,
                    defaults={
                        "name": name, "channel": channel,
                        "start_date": start_date, "end_date": end_date,
                        "spend": spend, "leads": leads, "bookings": bookings,
                        "cost_per_lead": cost_per_lead, "cost_per_booking": cost_per_booking,
                        "roi": roi, "status": status,
                    }
                )
                created["marketing_campaigns"] += 1
            except Exception as e:
                errors.append(f"MarketingCampaign row {row[:2]}: {e}")

    # Sheet "LocationDemandMonthly"
    if "LocationDemandMonthly" in wb.sheetnames:
        ws = wb["LocationDemandMonthly"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        today = date.today()
        for row in rows:
            if not row or row[0] is None:
                continue
            try:
                location = str(row[0]).strip()
                city = str(row[1]).strip() if len(row) > 1 and row[1] else ""
                year = int(row[2]) if len(row) > 2 and row[2] is not None else today.year
                month = int(row[3]) if len(row) > 3 and row[3] is not None else today.month
                enquiries = int(row[4]) if len(row) > 4 and row[4] is not None else 0
                bookings = int(row[5]) if len(row) > 5 and row[5] is not None else 0
                demand_score = Decimal(str(row[6])) if len(row) > 6 and row[6] is not None else Decimal(0)
                LocationDemandMonthly.objects.update_or_create(
                    organization=org, location=location, year=year, month=month,
                    defaults={"city": city, "enquiries": enquiries, "bookings": bookings, "demand_score": demand_score}
                )
                created["location_demand"] += 1
            except Exception as e:
                errors.append(f"LocationDemandMonthly row {row[:2]}: {e}")

    # Sheet "OrgKPI_Daily" (optional): date, total_units, revenue_booked, revenue_collected, outstanding, etc.
    if "OrgKPI_Daily" in wb.sheetnames:
        ws = wb["OrgKPI_Daily"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for row in rows:
            if not row or row[0] is None:
                continue
            try:
                d = row[0]
                if hasattr(d, "date"):
                    d = d.date()
                total_units = int(row[1]) if len(row) > 1 and row[1] is not None else 0
                revenue_booked = Decimal(str(row[2])) if len(row) > 2 and row[2] is not None else Decimal(0)
                revenue_collected = Decimal(str(row[3])) if len(row) > 3 and row[3] is not None else Decimal(0)
                outstanding = Decimal(str(row[4])) if len(row) > 4 and row[4] is not None else Decimal(0)
                OrgKPI_Daily.objects.update_or_create(
                    organization=org, date=d,
                    defaults={
                        "total_units": total_units, "revenue_booked": revenue_booked,
                        "revenue_collected": revenue_collected, "outstanding": outstanding,
                    }
                )
                created["org_kpi"] += 1
            except Exception as e:
                errors.append(f"OrgKPI_Daily row {row[:2]}: {e}")

    # Sheet "OrgMonthlySnapshot" (optional): year, month, revenue_booked, revenue_collected, cash_inflow, cash_outflow, etc.
    if "OrgMonthlySnapshot" in wb.sheetnames:
        ws = wb["OrgMonthlySnapshot"]
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        today = date.today()
        for row in rows:
            if not row or row[0] is None:
                continue
            try:
                # Expected columns:
                # 0: year (required)
                # 1: month (required, 1-12)
                # 2: total_units (optional, default 0)
                # 3: revenue_booked (optional, default 0)
                # 4: revenue_collected (optional, default 0)
                # 5: outstanding (optional, default 0)
                # 6: cash_inflow (optional, default 0)
                # 7: cash_outflow (optional, default 0)
                # 8: net_cashflow (optional, default 0)
                # 9: avg_satisfaction (optional, default 0)
                # 10: bookings_count (optional, default 0)
                # 11: avg_ticket_size (optional, default 0)
                year = int(row[0]) if len(row) > 0 and row[0] is not None else today.year
                month = int(row[1]) if len(row) > 1 and row[1] is not None else today.month
                
                # Validate month
                if month < 1 or month > 12:
                    errors.append(f"OrgMonthlySnapshot row {row[:2]}: Invalid month {month}. Must be 1-12.")
                    continue
                
                total_units = int(row[2]) if len(row) > 2 and row[2] is not None else 0
                revenue_booked = safe_decimal(row[3] if len(row) > 3 else None, "revenue_booked", row[:2])
                revenue_collected = safe_decimal(row[4] if len(row) > 4 else None, "revenue_collected", row[:2])
                outstanding = safe_decimal(row[5] if len(row) > 5 else None, "outstanding", row[:2])
                
                cash_inflow = safe_decimal(row[6] if len(row) > 6 else None, "cash_inflow", row[:2])
                cash_outflow = safe_decimal(row[7] if len(row) > 7 else None, "cash_outflow", row[:2])
                net_cashflow = safe_decimal(row[8] if len(row) > 8 else None, "net_cashflow", row[:2])
                
                avg_satisfaction = safe_decimal(row[9] if len(row) > 9 else None, "avg_satisfaction", row[:2])
                bookings_count = int(row[10]) if len(row) > 10 and row[10] is not None else 0
                avg_ticket_size = safe_decimal(row[11] if len(row) > 11 else None, "avg_ticket_size", row[:2])
                
                OrgMonthlySnapshot.objects.update_or_create(
                    organization=org, year=year, month=month,
                    defaults={
                        "total_units": total_units,
                        "revenue_booked": revenue_booked,
                        "revenue_collected": revenue_collected,
                        "outstanding": outstanding,
                        "cash_inflow": cash_inflow,
                        "cash_outflow": cash_outflow,
                        "net_cashflow": net_cashflow,
                        "avg_satisfaction": avg_satisfaction,
                        "bookings_count": bookings_count,
                        "avg_ticket_size": avg_ticket_size,
                    }
                )
                created["org_monthly_snapshot"] += 1
            except Exception as e:
                errors.append(f"OrgMonthlySnapshot row {row[:2]}: {e}")

    wb.close()
    return created, errors


class ExcelUploadAPIView(APIView):
    """
    POST /api/admin/upload-excel/
    Admin-only. Expects multipart: file (Excel), org_code (optional if Organization sheet has code).
    Creates/updates all org-scoped data including:
    Organization, Project, ProjectKPI_Daily, Unit, Customer, Booking,
    Vendor, VendorBill, VendorPayment, CashFlowEntry,
    Contractor, Milestone, DailyProgress, DelayPenalty,
    CustomerSatisfactionSurvey, Complaint,
    LegalCase, ComplianceItem, RERARegistration,
    OrgMonthlySnapshot, MarketingCampaign, LocationDemandMonthly, OrgKPI_Daily
    """
    permission_classes = [IsAuthenticated, IsAdminUser]
    parser_classes = [MultiPartParser, FormParser]

    def post(self, request):
        file = request.FILES.get("file") or request.data.get("file")
        if not file:
            return Response({"error": "No file provided. Use form field 'file'."}, status=400)
        if not file.name.endswith((".xlsx", ".xls")):
            return Response({"error": "File must be .xlsx or .xls"}, status=400)

        org_code = (request.data.get("org_code") or request.POST.get("org_code") or "").strip()
        org = Organization.objects.filter(code=org_code).first() if org_code else None

        created, errors = _parse_excel_upload(file, org)
        if errors and not created.get("organizations") and not org:
            return Response({"error": "Could not determine organization.", "details": errors}, status=400)
        return Response({
            "message": "Upload processed.",
            "created": created,
            "errors": errors[:20],
        }, status=200)
